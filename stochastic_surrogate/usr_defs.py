"""
Instantiate global variables of user definitions made in user_input.xlsx
"""

import os as _os
import pandas as _pd
import numpy as _np
from openpyxl import load_workbook
from config import *


def assign_calib_ranges(direct_par_df, indirect_par_df, recalc_par_df):
    """Parse user calibration ranges for parameters

    :param pd.DataFrame direct_par_df: direct calibration parameters from user-input.xlsx
    :param pd.DataFrame indirect_par_df: indirect calibration parameters from user-input.xlsx
    :param pd.DataFrame recalc_par_df: recalculation parameters from user-input.xlsx
    :return:
    """
    global CALIB_PAR_SET  # dict for calibration optimization parameters and ranges
    global CALIB_ID_PAR_SET  # dict for indirect calibration parameters
    CALIB_PAR_SET = {}
    CALIB_ID_PAR_SET = {}

    dir_par_dict = dict(zip(direct_par_df[0].to_list(), direct_par_df[1].to_list()))
    for par, bounds in dir_par_dict.items():
        if not (("TELEMAC" or "GAIA") in par):
            CALIB_PAR_SET.update({par: {"bounds": str2seq(bounds),
                                        "distribution": None}})

    indir_par_dict = dict(zip(indirect_par_df[0].to_list(), indirect_par_df[1].to_list()))
    for par, bounds in indir_par_dict.items():
        CALIB_ID_PAR_SET.update({par: {"classes": str2seq(bounds),
                                       "distribution": None}})
        if not (("TELEMAC" or "GAIA") in par):
            # erase CALIB_ID_PAR_SET in the last step, if user did not enable (OK because inexpensive)
            CALIB_ID_PAR_SET = None

    recalc_par_dict = dict(zip(recalc_par_df[0].to_list(), recalc_par_df[1].to_list()))
    for par, bounds in recalc_par_dict.items():
        # loop not really needed but implemented for potential developments
        if not(par in dir_par_dict) and bounds:
            # overwrite or add recalculation parameter in CALIB_PAR_SET dict
            # here: bounds is a user-defined boolean
            CALIB_PAR_SET.update({par: {"bounds": (CALIB_ID_PAR_SET["Multiplier range"]["classes"][0],
                                                   CALIB_ID_PAR_SET["Multiplier range"]["classes"][1]),
                                        "distribution": None}})

    print(" * received direct calibration parameters: %s" % ", ".join(list(CALIB_PAR_SET.keys())))
    if CALIB_ID_PAR_SET:
        print(" * received indirect calibration parameter: %s" % ", ".join(list(CALIB_ID_PAR_SET.keys())))


def check_user_input():
    """Check if global variables are correctly assigned"""
    print(" * verifying directories...")
    if not (_os.path.isdir(SIM_DIR)):
        print("ERROR: Cannot find %s - please double-check input XLSX (cell B8).")
        raise NotADirectoryError
    if not (_os.path.isfile(SIM_DIR + "/%s" % TM_CAS)):
        print("ERROR: The TELEMAC steering file %s does not exist." % str(SIM_DIR + "/%s" % TM_CAS))
        raise FileNotFoundError
    if GAIA_CAS:
        if not (_os.path.isfile(SIM_DIR + "/%s" % GAIA_CAS)):
            print("ERROR: The GAIA steering file %s does not exist." % str(SIM_DIR + "/%s" % GAIA_CAS))
            raise FileNotFoundError
    if not (_os.path.isfile(CALIB_PTS)):
        print("ERROR: The Calibration CSV file %s does not exist." % str(CALIB_PTS))
        raise FileNotFoundError
    if not (_os.path.isdir(RESULTS_DIR)):
        try:
            _os.mkdir(RESULTS_DIR)
        except PermissionError:
            print("ERROR: Cannot write to %s (check user rights/path consistency)" % RESULTS_DIR)
            raise PermissionError
        except NotADirectoryError:
            print("ERROR: %s is not a directory - adapt simulation directory (B8)" % RESULTS_DIR)
            raise NotADirectoryError
    if MC_SAMPLES < (AL_SAMPLES + IT_LIMIT):
        print("ERROR: MC_SAMPLES < (AL_SAMPLES + IT_LIMIT)!")
        raise ValueError


def load_input_defs():
    """loads provided input file name as dictionary

    Returns:
        (dict): user input of input.xlsx (or custom file, if provided)
    """
    print(" * loading %s" % INPUT_XLSX_NAME)
    return {
        "tm pars": read_wb_range(INPUT_XLSX_NAME, TM_RANGE),
        "al pars": read_wb_range(INPUT_XLSX_NAME, AL_RANGE),
        "direct priors": read_wb_range(INPUT_XLSX_NAME, PRIOR_DIR_RANGE),
        "indirect priors": read_wb_range(INPUT_XLSX_NAME, PRIOR_INDIR_RANGE),
        "recalculation priors": read_wb_range(INPUT_XLSX_NAME, PRIOR_REC_RANGE),
    }


def str2seq(list_like_string, separator=",", return_type="tuple"):
    """Convert a list-like string into a tuple or list based on a separator such as comma or semi-column

    :param str list_like_string: string to convert
    :param str separator: separator to use
    :param str return_type: defines if a list or tuple is returned (default: tuple)
    :return: list or tuple
    """
    seq = []
    for number in list_like_string.split(separator):
        try:
            seq.append(float(number))
        except ValueError:
            print("WARNING: Could not interpret user parameter value range definition (%s)" % number)
            print("         This Warning will probably cause an ERROR later in the script.")
    if "tuple" in return_type:
        return tuple(seq)
    else:
        return seq


def rewrite_globals(file_name="user-input.xlsx"):
    """rewrite globals from config

    Args:
        file_name (str): name of input file (default is user-input.xlsx)

    Returns:
        (dict): user input of input.xlsx (or custom file, if provided)
    """
    # instantiate global variables to modify
    global INPUT_XLSX_NAME  # str of path to user-input.xlsx including filename
    global CALIB_PTS  # numpy array to be loaded from calibration_points file
    global CALIB_TARGET # str of calibration target nature (e.g. topographic change)
    global AL_STRATEGY  # str for active learning strategy
    global IT_LIMIT  # int limit for Bayesian iterations
    global MC_SAMPLES  # int for Monte Carlo samples
    global MC_SAMPLES_AL  # int for Monte Carlo samples
    global N_CPUS  # int number of CPUs to use for Telemac models
    global AL_SAMPLES  # int for no. of active learning sampling size
    global TM_CAS
    global GAIA_CAS
    global SIM_DIR
    global RESULTS_DIR

    # update input xlsx file name globally and load user definitions
    INPUT_XLSX_NAME = file_name
    user_defs = load_input_defs()  # dict

    print(" * assigning user-defined variables...")
    # assign direct, indirect, and recalculation parameters
    assign_calib_ranges(
        direct_par_df=user_defs["direct priors"],
        indirect_par_df=user_defs["indirect priors"],
        recalc_par_df=user_defs["recalculation priors"]
    )

    # update global variables with user definitions
    CALIB_PTS = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("calib\_pts"), 1].values[0]
    CALIB_TARGET = TM_TRANSLATOR[user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("calib\_target"), 1].values[0]]

    AL_STRATEGY = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("strategy"), 1].values[0]
    IT_LIMIT = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("it\_limit"), 1].values[0]
    AL_SAMPLES = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("al\_samples"), 1].values[0]
    MC_SAMPLES = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("mc\_samples\)"), 1].values[0]
    MC_SAMPLES_AL = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("mc\_samples\_al"), 1].values[0]

    TM_CAS = user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("TELEMAC"), 1].values[0]
    GAIA_CAS = user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("Gaia"), 1].values[0]
    SIM_DIR = r"" + user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("Simulation"), 1].values[0]
    N_CPUS = user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("CPU"), 1].values[0]
    RESULTS_DIR = SIM_DIR + "opt-results/"

    check_user_input()


def read_wb_range(file_name, read_range, sheet_name="MAIN"):
    """Read a certain range of a workbook only with openpyxl

    :param str file_name: full path and name of input file
    :param str read_range: letter-number read range in workbook (e.g. "A2:B4")
    :param str sheet_name: name of the sheet to read (default is MAIN from user-inpux.xlsx)
    :return pd.DataFrame: xlsx contents in the defined range
    """
    ws = load_workbook(filename=file_name, read_only=True, data_only=True)[sheet_name]
    # Read the cell values into a list of lists
    data_rows = []
    for row in ws[read_range]:
        data_rows.append([cell.value for cell in row])
    return _pd.DataFrame(data_rows)

