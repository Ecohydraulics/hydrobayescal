"""
Instantiate global variables of user definitions made in user_input.xlsx
"""

import os as _os
import pandas as _pd
import numpy as _np
from openpyxl import load_workbook
from basic_functions import *
import random as _rnd


class UserDefs:
    def __init__(self, input_worbook_name="user-input.xlsx", *args, **kwargs):
        self.input_xlsx_name = input_worbook_name
        print(
            "Using %s to read user settings. To change the user settings, run OBJECT(bal_gpe4telemac).write_global_settings('path/to/workbook.xlsx')" % self.input_xlsx_name)

        self.file_write_dir = ""

        # initialize capital letter class variables that are defined through the user XLSX file
        self.CALIB_PAR_SET = {}  # dict for direct calibration optimization parameters
        self.CALIB_PTS = None  # numpy array to be loaded from calibration_points file
        self.CALIB_TARGET = str()  # str of calibration target nature (e.g. topographic change)
        self.init_runs = int()  # int limit for initial full-complexity runs
        self.IT_LIMIT = int()  # int limit for Bayesian iterations
        self.MC_SAMPLES = int()  # int for Monte Carlo samples
        self.MC_SAMPLES_AL = int()  # int for Monte Carlo samples for active learning
        self.N_CPUS = int()  # int number of CPUs to use for Telemac models
        self.n_calib_pars = int()  # int number of calibration parameters
        self.AL_SAMPLES = int()  # int for no. of active learning sampling size
        self.AL_STRATEGY = str()  # str for active learning strategy
        self.TM_CAS = str()
        self.tm_config = str()  # telemac config file
        self.GAIA_CAS = str()
        self.RESULTS_DIR = "../results"  # relative path for results
        self.SIM_DIR = str()  # relative path for simulations
        self.BME = None
        self.RE = None
        self.al_BME = None
        self.al_RE = None

    def assign_calib_ranges(self, direct_par_df, vector_par_df, recalc_par_df):
        """Parse user calibration ranges for parameters

        :param pd.DataFrame direct_par_df: direct calibration parameters from user-input.xlsx
        :param pd.DataFrame vector_par_df: vector calibration parameters from user-input.xlsx
        :param pd.DataFrame recalc_par_df: recalculation parameters from user-input.xlsx
        :return: None
        """
        # add scalar calibration parameters to CALIB_PAR_SET
        dir_par_dict = dict(zip(direct_par_df[0].to_list(), direct_par_df[1].to_list()))
        for par, bounds in dir_par_dict.items():
            if not (("TELEMAC" or "GAIA") in par):
                self.CALIB_PAR_SET.update({par: {"bounds": str2seq(bounds),
                                                 "initial val": _np.mean(str2seq(bounds)),
                                                 "recalc par": None}})

        # add vector calibration parameters to CALIB_PAR_SET and check for recalculation parameters
        vec_par_dict = dict(zip(vector_par_df[0].to_list(), vector_par_df[1].to_list()))
        recalc_par_dict = dict(zip(recalc_par_df[0].to_list(), recalc_par_df[1].to_list()))
        for par, init_list in vec_par_dict.items():
            if not (("TELEMAC" or "GAIA" or "Multiplier") in str(par)):
                self.CALIB_PAR_SET.update({par: {"bounds": (str2seq(vec_par_dict["Multiplier range"])),
                                                 "initial val": str2seq(init_list),
                                                 "recalc par": None}})
                if par in RECALC_PARS.keys():
                    # check if parameter is a recalculation parameter (if yes -> check for user input FALSE or TRUE)
                    try:
                        if bool(recalc_par_dict[RECALC_PARS[par]]):
                            self.CALIB_PAR_SET[par]["recalc par"] = RECALC_PARS[par]
                    except KeyError:
                        print("! Warning: found recalcution parameter %s that is not defined in config.py (skipping...")
        self.n_calib_pars = len(self.CALIB_PAR_SET)
        print(" * received the following calibration parameters: %s" % ", ".join(list(self.CALIB_PAR_SET.keys())))

    def check_user_input(self):
        """Check if global variables are correctly assigned"""
        print(" * verifying directories...")
        if not (_os.path.isdir(self.SIM_DIR)):
            print("ERROR: Cannot find %s - please double-check input XLSX (cell B8).")
            raise NotADirectoryError
        if not (_os.path.isfile(self.SIM_DIR + "/%s" % self.TM_CAS)):
            print("ERROR: The TELEMAC steering file %s does not exist." % str(self.SIM_DIR + "/%s" % self.TM_CAS))
            raise FileNotFoundError
        if self.GAIA_CAS:
            if not (_os.path.isfile(self.SIM_DIR + "/%s" % self.GAIA_CAS)):
                print("ERROR: The GAIA steering file %s does not exist." % str(self.SIM_DIR + "/%s" % self.GAIA_CAS))
                raise FileNotFoundError
        if not (_os.path.isfile(self.CALIB_PTS)):
            print("ERROR: The Calibration CSV file %s does not exist." % str(self.CALIB_PTS))
            raise FileNotFoundError
        if not (_os.path.isdir(self.RESULTS_DIR)):
            try:
                _os.mkdir(self.RESULTS_DIR)
            except PermissionError:
                print("ERROR: Cannot write to %s (check user rights/path consistency)" % self.RESULTS_DIR)
                raise PermissionError
            except NotADirectoryError:
                print("ERROR: %s is not a directory - adapt simulation directory (B8)" % self.RESULTS_DIR)
                raise NotADirectoryError
        if self.MC_SAMPLES < (self.AL_SAMPLES + self.IT_LIMIT):
            print("ERROR: MC_SAMPLES < (AL_SAMPLES + IT_LIMIT)!")
            raise ValueError
        try:
            self.file_write_dir = r"" + self.RESULTS_DIR + "stochastic-calib-processID%s/" % str(_rnd.randint(1000,9999))
            _os.mkdir(self.file_write_dir)
            print(" * intermediate calibration results will be written to %s" % self.file_write_dir)
        except PermissionError:
            print("ERROR: Cannot write to %s (check user rights/path consistency)" % self.RESULTS_DIR)
            raise PermissionError
        except NotADirectoryError:
            print("ERROR: %s is not a directory - adapt simulation directory (B8)" % self.RESULTS_DIR)
            raise NotADirectoryError

    def load_input_defs(self):
        """loads provided input file name as dictionary

        Returns:
            (dict): user input of input.xlsx (or custom file, if provided)
        """
        print(" * loading %s" % self.input_xlsx_name)
        return {
            "tm pars": self.read_wb_range(TM_RANGE),
            "al pars": self.read_wb_range(AL_RANGE),
            "direct priors": self.read_wb_range(PRIOR_SCA_RANGE),
            "vector priors": self.read_wb_range(PRIOR_VEC_RANGE),
            "recalculation priors": self.read_wb_range(PRIOR_REC_RANGE),
        }

    def write_global_settings(self, file_name=None):
        """rewrite globals from config

        Args:
            file_name (str): name of input file (default is user-input.xlsx)

        Returns:
            (dict): user input of input.xlsx (or custom file, if provided)
        """

        # update input xlsx file name globally and load user definitions
        if file_name:
            self.input_xlsx_name = file_name
        user_defs = self.load_input_defs()  # dict

        print(" * assigning user-defined variables...")
        # assign direct, vector, and recalculation parameters
        self.assign_calib_ranges(
            direct_par_df=user_defs["direct priors"],
            vector_par_df=user_defs["vector priors"],
            recalc_par_df=user_defs["recalculation priors"]
        )

        # update global variables with user definitions
        self.CALIB_PTS = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("calib\_pts"), 1].values[0]
        self.CALIB_TARGET = TM_TRANSLATOR[
            user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("calib\_target"), 1].values[0]]

        self.AL_STRATEGY = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("strategy"), 1].values[0]
        self.init_runs = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("init\_runs"), 1].values[0]
        self.IT_LIMIT = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("it\_limit"), 1].values[0]
        self.AL_SAMPLES = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("al\_samples"), 1].values[0]
        self.MC_SAMPLES = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("mc\_samples\)"), 1].values[0]
        self.MC_SAMPLES_AL = user_defs["al pars"].loc[user_defs["al pars"][0].str.contains("mc\_samples\_al"), 1].values[0]

        self.TM_CAS = user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("TELEMAC steering"), 1].values[0]
        self.tm_config = user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("TELEMAC config"), 1].values[0]
        self.GAIA_CAS = user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("Gaia"), 1].values[0]
        self.SIM_DIR = r"" + user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("Simulation"), 1].values[0]
        self.N_CPUS = user_defs["tm pars"].loc[user_defs["tm pars"][0].str.contains("CPU"), 1].values[0]
        self.RESULTS_DIR = self.SIM_DIR + "opt-results/"

        # global surrogate and active learning variables
        self.BME = _np.zeros((self.IT_LIMIT, 1))
        self.RE = _np.zeros((self.IT_LIMIT, 1))
        self.al_BME = _np.zeros((self.AL_SAMPLES, 1))
        self.al_RE = _np.zeros((self.AL_SAMPLES, 1))

        self.check_user_input()

    def read_wb_range(self, read_range, sheet_name="MAIN"):
        """Read a certain range of a workbook only with openpyxl

        :param str read_range: letter-number read range in workbook (e.g. "A2:B4")
        :param str sheet_name: name of the sheet to read (default is MAIN from user-inpux.xlsx)
        :return pd.DataFrame: xlsx contents in the defined range
        """
        ws = load_workbook(filename=self.input_xlsx_name, read_only=True, data_only=True)[sheet_name]
        # Read the cell values into a list of lists
        data_rows = []
        for row in ws[read_range]:
            data_rows.append([cell.value for cell in row])
        return _pd.DataFrame(data_rows)
